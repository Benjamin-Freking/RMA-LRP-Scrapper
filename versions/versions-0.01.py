import datetime
import time
import urllib.request
import os
import zipfile
import csv
import pandas as pd
from openpyxl import load_workbook

# Constants
TARGET_STATE_CODE = '19'

MAX_RETRIES = 3
SAVE_DIRECTORY = "var/"
EXCEL_FILE_PATH = "LRP_Swine.xlsx"

dev_mode = input('Press ENTER to Skip or type "yes" to Enter Dev Mode: ')

if dev_mode.lower() == 'yes':
    OVERWRITE_DATE = input("Enter the OVERWRITE_DATE (YYYYMMDD format), or press Enter to use the default: ")
    if not OVERWRITE_DATE:
        OVERWRITE_DATE = None
else:
    OVERWRITE_DATE = None

NEW_COMMODITY_DIRECTORY = {
    '0801': {
        'directory_name': 'FeederCattle',
        'sub_sheets': {
            '809': '809_Sheet',
            '810': '810_Sheet',
            '811': '811_Sheet',
            '812': '812_Sheet',
        }
    },
    '0815': {
        'directory_name': 'FeederCattle',
        'sub_sheets': {
            '997': '997_Sheet',
            '821': '821_Sheet',
        }
    },
    '0802': 'FedCattle',
    '0802': {
        'directory_name': 'FedCattle',
        'sub_sheets': {
            '820': '820_Sheet'
        }
    },
}

# Define a function to process each commodity and return sheet name and DataFrame
def commodity_sheet_build(csv_data, key, sub_key, sub_value):
    sheet_name = sub_value
    # Specify the target values
    target_commodity_code = key
    print(f"Processing {sheet_name} ({target_commodity_code})")

    # Filter rows where 'Commodity Code' is the target and 'State Code' is the target
    matching_rows = []
    for row in csv_data:
        commodity_code = row.get('Commodity Code', '')
        state_code = row.get('State Code', '')
        type_code = row.get('Type Code', '')

        if commodity_code == target_commodity_code and state_code == TARGET_STATE_CODE and type_code==sub_key:
            matching_rows.append(row)

    df = pd.DataFrame(matching_rows)

    df = df.sort_values(by=[df.columns[11], df.columns[12]])
    df = df.reset_index(drop=True)

    columns_to_drop = list(range(1, 4)) + list(range(5, 11)) + list(range(13, 21)) + list(
        range(28, 34))
    df = df.drop(df.columns[columns_to_drop], axis=1)

    print(f"Succesfully Processed {sheet_name} ({target_commodity_code})")
    return sheet_name, df

def download_and_extract_file(url, save_directory, max_retries):
    retry_count = 0
    while retry_count < max_retries:
        try:
            if not os.path.exists(save_directory):
                os.makedirs(save_directory)
            # Download the file
            urllib.request.urlretrieve(url, os.path.join(save_directory, filename))
            print(f"File '{filename}' downloaded successfully to '{save_directory}'")

            # Extract the ZIP file
            with zipfile.ZipFile(os.path.join(save_directory, filename), 'r') as zip_ref:
                file_list = zip_ref.namelist()

                # Check the number of files inside the ZIP
                num_files = len(file_list)
                print(f"Number of files inside the ZIP: {num_files}")

                # If there are two files, look for 'LrpRate' in the name
                for file_name in file_list:
                    if 'LrpRate' in file_name:
                        # Extract the specific file
                        zip_ref.extract(file_name, path=save_directory)

                        # Open and read the CSV file
                        csv_file_path = os.path.join(save_directory, file_name)
                        print(save_directory + file_name)
                        with open(csv_file_path, 'r') as csv_file:
                            csv_data = list(csv.DictReader(csv_file, delimiter='|'))

                        commodity_dfs = []

                        for key in NEW_COMMODITY_DIRECTORY:
                            print(key)
                            if 'sub_sheets' in NEW_COMMODITY_DIRECTORY[key]:
                                for sub_key, sub_value in NEW_COMMODITY_DIRECTORY[key]['sub_sheets'].items():
                                    print(f"Sub-sheet Key: {sub_key}, Sub-sheet Value: {sub_value}")
                                    result = commodity_sheet_build(csv_data, key, sub_key, sub_value)
                                    if result is not None:
                                        sheet_name, df = result
                                        commodity_dfs.append((sheet_name, df))

                        return commodity_dfs

            break  # Exit the loop if the download is successful
        except Exception as e:
            print(f"Error downloading or checking file: {e}")
            retry_count += 1
            if retry_count < max_retries:
                print(f"Retrying in 5 minutes...")
                time.sleep(300)  # Wait for 5 minutes before retrying

    if retry_count >= max_retries:
        print(f"Maximum retries reached. File '{filename}' not downloaded.")

# Base URL and year
base_url = "https://pubfs-rma.fpac.usda.gov/pub/References/adm_livestock/"
year = datetime.datetime.now().year + 1

# Get the current date and subtract one day
current_date = datetime.datetime.now()
current_date_str = current_date.strftime("%Y%m%d")
if OVERWRITE_DATE:
    current_date_str = OVERWRITE_DATE

# Construct the filename and URL
filename = f"{year}_ADMLivestockLrp_Daily_{current_date_str}.zip"
url = f"{base_url}{year}/{filename}"
print(f"Gathering RMA Data for Date: {current_date_str}")
print(f"Gathering RMA Data from URL: {url}")


# Create a list to store DataFrames for each commodity


# Download and extract the file
commodity_dfs = download_and_extract_file(url, SAVE_DIRECTORY, MAX_RETRIES)
try:
    print(len(commodity_dfs))
except:
    print(f'No Data Pulled for {current_date_str}')

# Save the changes to the Excel file
wb = load_workbook(EXCEL_FILE_PATH)

for sheet_name, df in commodity_dfs:
    print(sheet_name)
    # Get the sheet or create it if it doesn't exist
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(title=sheet_name)

    # Clear the content of the sheet
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.value = None

    def calculate_new_column_value(row):
        feed_value = pd.to_numeric(row[8], errors='coerce')  # Assuming feed value is in column 8
        condition_value = pd.to_numeric(row[10], errors='coerce')  # Assuming condition value is in column 10
        print(feed_value)
        print(condition_value)
        if 0.95 <= feed_value <= 1.00:
            return condition_value * (1 - 0.35)
        elif 0.90 <= feed_value <= 0.9499:
            return condition_value * (1 - 0.40)
        elif 0.85 <= feed_value <= 0.8999:
            return condition_value * (1 - 0.45)
        elif 0.80 <= feed_value <= 0.8499:
            return condition_value * (1 - 0.50)
        elif 0.70 <= feed_value <= 0.7999:
            return condition_value * (1 - 0.55)
        else:
            return 0.0  # Default value when none of the conditions are met

    # Apply the custom function to create the 'NewColumn'
    df['NewColumn'] = df.apply(calculate_new_column_value, axis=1)
    print(df)
    # Iterate through the sorted DataFrame and paste data into the Excel sheet
    for index, row in df.iterrows():
        for col_idx, value in enumerate(row, 1):
            # Start pasting at row 2 (skip the header row)
            cell = sheet.cell(row=index + 2, column=col_idx, value=value)

# Save the changes to the Excel file
wb.save(EXCEL_FILE_PATH)


