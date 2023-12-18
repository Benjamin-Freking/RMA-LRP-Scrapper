import datetime
import urllib.request
import os
import zipfile
import csv
import pandas as pd
from openpyxl import load_workbook
import time
import sys

# Constants
TARGET_STATE_CODE = '19'
MAX_RETRIES = 24
SAVE_DIRECTORY = "var/"
EXCEL_FILE_PATH = "LRP_Swine.xlsx"

# Commodity directory configuration
NEW_COMMODITY_DIRECTORY = {
    '0801': {
        'directory_name': 'FeederCattle',
        'sub_sheets': {
            '809': '809_Sheet',
            '810': '810_Sheet',
            '811': '811_Sheet',
            '812': '812_Sheet',
        }
    },
    '0815': {
        'directory_name': 'FeederCattle',
        'sub_sheets': {
            '997': '997_Sheet',
            '821': '821_Sheet',
        }
    },
    '0802': {
        'directory_name': 'FedCattle',
        'sub_sheets': {'820': '820_Sheet'}
    },
}

# Developer mode setting
dev_mode = input('Press ENTER to Skip or type "yes" to Enter Dev Mode: ')
OVERWRITE_DATE = None
if dev_mode.lower() == 'yes':
    OVERWRITE_DATE = input("Enter the OVERWRITE_DATE (YYYYMMDD format), or press Enter to use the default: ")

def commodity_sheet_build(csv_data, key, sub_key, sub_value):
    """
    Process each commodity and return sheet name and DataFrame.

    Args:
        csv_data (list): List of dictionaries representing CSV data.
        key (str): Commodity code.
        sub_key (str): Sub commodity code.
        sub_value (str): Sheet name.

    Returns:
        tuple: Sheet name and DataFrame.
    """
    sheet_name = sub_value
    print(f"Processing {sheet_name} ({key})")

    # Filter rows with specified commodity and state codes
    matching_rows = [
        row for row in csv_data
        if row.get('Commodity Code', '') == key and
           row.get('State Code', '') == TARGET_STATE_CODE and
           row.get('Type Code', '') == sub_key
    ]

    # Create DataFrame and process data
    df = pd.DataFrame(matching_rows)
    df = df.sort_values(by=[df.columns[11], df.columns[12]]).reset_index(drop=True)
    columns_to_drop = list(range(1, 4)) + list(range(5, 11)) + list(range(13, 21)) + list(range(28, 34))
    df = df.drop(df.columns[columns_to_drop], axis=1)

    print(f"Succesfully Processed {sheet_name} ({key})")
    return sheet_name, df

def download_and_extract_file(url, save_directory, max_retries):
    """
    Download and extract a file from a given URL.

    Args:
        url (str): URL to download the file from.
        save_directory (str): Directory to save the downloaded file.
        max_retries (int): Maximum number of retry attempts.

    Returns:
        list: List of tuples containing sheet names and corresponding DataFrames.
    """
    retry_count = 0
    while retry_count < max_retries:
        try:
            filename = os.path.basename(url)
            if not os.path.exists(save_directory):
                os.makedirs(save_directory)
            urllib.request.urlretrieve(url, os.path.join(save_directory, filename))
            print(f"File '{filename}' downloaded successfully to '{save_directory}'")

            with zipfile.ZipFile(os.path.join(save_directory, filename), 'r') as zip_ref:
                zip_ref.extractall(save_directory)

                for file_name in zip_ref.namelist():
                    if 'LrpRate' in file_name:
                        csv_file_path = os.path.join(save_directory, file_name)
                        with open(csv_file_path, 'r') as csv_file:
                            csv_data = list(csv.DictReader(csv_file, delimiter='|'))
                            commodity_dfs = [
                                commodity_sheet_build(csv_data, key, sub_key, sub_value)
                                for key, value in NEW_COMMODITY_DIRECTORY.items()
                                if 'sub_sheets' in value
                                for sub_key, sub_value in value['sub_sheets'].items()
                            ]
                            return commodity_dfs

            break
        except Exception as e:
            print(f"Error downloading or processing file: {e}")
            retry_count += 1
            if retry_count < max_retries:
                print("Retrying in 5 minutes...")
                time.sleep(300)

    if retry_count >= max_retries:
        print(f"Maximum retries reached. File '{filename}' not downloaded.")

# URL construction
base_url = "https://pubfs-rma.fpac.usda.gov/pub/References/adm_livestock/"
year = datetime.datetime.now().year + 1
current_date_str = datetime.datetime.now().strftime("%Y%m%d") if not OVERWRITE_DATE else OVERWRITE_DATE
filename = f"{year}_ADMLivestockLrp_Daily_{current_date_str}.zip"
url = f"{base_url}{year}/{filename}"
print(f"Gathering RMA Data for Date: {current_date_str}")
print(f"Gathering RMA Data from URL: {url}")

def print_countdown_timer(seconds):
    """
    Print a countdown timer.

    Args:
        seconds (int): Number of seconds for the countdown.
    """
    for remaining in range(seconds, 0, -1):
        minutes, seconds = divmod(remaining, 60)
        sys.stdout.write(f"\rCountdown Timer: {minutes}:{seconds:02}")
        sys.stdout.flush()
        time.sleep(1)

# Main download and processing loop
try:
    while True:
        try:
            commodity_dfs = download_and_extract_file(url, SAVE_DIRECTORY, MAX_RETRIES)
            if commodity_dfs:
                print(f"Number of DataFrames processed: {len(commodity_dfs)}")
                break
        except Exception as e:
            print(f"No Data Pulled for {datetime.datetime.now().strftime('%Y-%m-%d at %H:%M:%S')}")
            print('\nRestarting Program - Press "Ctrl+C" to exit the program if needed.')
            print_countdown_timer(300)  # 5-minute countdown
except KeyboardInterrupt:
    print('Program terminated by user.')

# Saving to Excel file
try:
    wb = load_workbook(EXCEL_FILE_PATH)
    for sheet_name, df in commodity_dfs:
        print(f"Updating Sheet: {sheet_name}")
        sheet = wb.get_sheet_by_name(sheet_name) if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
        sheet.delete_rows(2, sheet.max_row)

        def calculate_new_column_value(row):
            feed_value = pd.to_numeric(row[8], errors='coerce')  # Assuming feed value is in column 8
            condition_value = pd.to_numeric(row[10], errors='coerce')  # Assuming condition value is in column 10
            if 0.95 <= feed_value <= 1.00:
                return condition_value * (1 - 0.35)
            elif 0.90 <= feed_value <= 0.9499:
                return condition_value * (1 - 0.40)
            elif 0.85 <= feed_value <= 0.8999:
                return condition_value * (1 - 0.45)
            elif 0.80 <= feed_value <= 0.8499:
                return condition_value * (1 - 0.50)
            elif 0.70 <= feed_value <= 0.7999:
                return condition_value * (1 - 0.55)
            else:
                return 0.0  # Default value when none of the conditions are met


        # Apply the custom function to create the 'NewColumn'
        df['NewColumn'] = df.apply(calculate_new_column_value, axis=1)
        print(f"Updated Producer Premium for: {sheet_name}")

        # Iterate through the sorted DataFrame and paste data into the Excel sheet
        for index, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                # Start pasting at row 2 (skip the header row)
                cell = sheet.cell(row=index + 2, column=col_idx)
                cell.value = value

    wb.save(EXCEL_FILE_PATH)
    print("Excel Workbook saved.")
except TypeError as te:
    if "'NoneType' object is not iterable" in str(te):
        print("RMA Datapull Empty - Failed to gather Dataframes and update Excel Sheet")
    else:
        print(f"An error occurred: {te}")
except Exception as e:
    print(f"An error occurred: {e}")
# Save the changes to the Excel file
